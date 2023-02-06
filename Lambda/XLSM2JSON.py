import boto3
import json
import urllib.parse
from collections import OrderedDict
from itertools import islice
from openpyxl import load_workbook

s3_client = boto3.client("s3")
S3_BUCKET_NAME = 'skillprofilebucket'

dynamodb = boto3.resource('dynamodb')
table = dynamodb.Table('SkillProfile')

def lambda_handler(event, context):
    
    #  PART 1: Connect to S3 and download the EXCEL file (to /tmp/ in Lambda)
   
    object_key = urllib.parse.unquote_plus(event['Records'][0]['s3']['object']['key'], encoding='utf-8') 
    
    file_content = s3_client.download_file(S3_BUCKET_NAME, object_key,'/tmp/data.xlsm')
    
    #  PART 2: Parse data from EXCEL file to Python list using openpyxl

    wb = load_workbook('/tmp/data.xlsm')
    
    sheet1 = wb['Mapping SSP']  #Sheet with Profiles' info
    
    profile_list = []
    diccionario = {}
    
    for row in islice(sheet1.values, 3, (sheet1.max_row)-1):
        profile = OrderedDict()
        profile['ProfileID'] = row[1]
        profile['ID'] = row[1]
        profile['Name'] = row[2]
        profile['Family'] = row[16]
        profile['Cluster'] = row[17]
        profile['Description'] = row[4]
        profile_list.append(profile)
        diccionario[row[2].upper()]=row[1]
    
    
    sheet2 = wb['Skill Profiles']   #Sheet with Skills' info
    
    skill_list = []
    
    for row in islice(sheet2.values, 11, (sheet2.max_row)-1):
        skill = OrderedDict()
        skill['ProfileID'] = diccionario[row[4].upper()]
        skill['ID'] = row[5]
        skill['Category'] = row[6]
        skill['SkillName'] = row[9]
        skill['Description'] = row[10]
        Levels = OrderedDict()
        Levels['Junior'] = row[11]
        Levels['Specialist'] = row[12]
        Levels['Expert'] = row[13]
        Levels['Senior'] = row[14]
        Levels['Principal'] = row[15]
        skill['Levels']=Levels
        skill_list.append(skill)
    
    #  PART 3: Connect to DynamoDB and upload the data to the table
    
    for profile in profile_list:
        table.put_item(Item=profile)
    for skill in skill_list:
        table.put_item(Item=skill)

